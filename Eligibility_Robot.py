
import smtplib
from pywinauto import keyboard
from selenium import webdriver
from datetime import datetime as dt
import openpyxl, time, re, os, datetime
from selenium.webdriver.common.by import By
from pywinauto.application import Application
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


url = r'https://mytools.gatewayedi.com/logon?ReturnUrl=%2f'
driver = webdriver.Chrome()
driver.maximize_window()
driver.get(url)
time.sleep(2)

username = '#####'
password = '#########'

find_user = driver.find_element(By.ID,'UserName')
find_user.send_keys(username)
print('username entered successfully')
time.sleep(1.5)

find_pass = driver.find_element(By.ID,'Password')
find_pass.send_keys(password)
print('password entered successfully')
time.sleep(1.5)

logIn = driver.find_element(By.ID,'login-button').click()
WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="NavCtrl_navManagePatients"]')))
time.sleep(1.5)
print('login successfull')

manage_patients = driver.find_element(By.XPATH,'//*[@id="NavCtrl_navManagePatients"]')
manage_patients.click()
print('manage patients click successfull')
time.sleep(1.5)

FFM_Bot_Output_File_Path = r"H:\AR PRODUCTION REPORTS\Business Intelligence\Eligibility and Benefits BOT Output.xlsx"
workbook = openpyxl.load_workbook(FFM_Bot_Output_File_Path)
sheet = workbook['Data']

pattern = r'\$?\d+(\.\d+)?%?'

folder_path = r"H:\AR PRODUCTION REPORTS\Business Intelligence\SOP\Fulshear Family Medicine_881\PDF Downloads"
current_Datetime = dt.now()
today_Date = current_Datetime.strftime("%m-%d-%Y")
today_folder_path = os.path.join(folder_path, today_Date)
if not os.path.exists(today_folder_path):
    os.mkdir(today_folder_path)
    print(f"Created folder: {today_Date}")

today = datetime.date.today()
today_date = today.strftime("%m/%d/%Y")

def BCBS():
    try:
        plan_date = driver.find_element(By.XPATH,'//*[@id="BasicProfile"]/dl/dt[contains(text(),"Plan Date")]//following::dd').text
        list1 = plan_date.split('-')
        Policy_Effective_date = list1[0] if len(list1) > 0 else None
        Policy_Term_date = list1[1] if len(list1) > 1 else None
        row[18].value = Policy_Effective_date
        row[19].value = Policy_Term_date
        workbook.save(FFM_Bot_Output_File_Path)
        print('Plan Dates Done')
    except:
        pass

    InsuranceType = driver.find_element(By.XPATH,'//a[text()="Active Coverage"]/following::table[1]//tr[2]/td[3]').text
    row[21].value = InsuranceType
    if '(HMO)' in InsuranceType:
        row[20].value = 'Yes'
    else:
        row[20].value = 'No'
    try:
        PCP_Name = driver.find_element(By.XPATH,'//a[text()="Primary Care Provider"]/following::table[1]//tr[3]/td/dl/dt[text()="Primary Care Provider"]/following-sibling::dd').text
    except:
        try:
            PCP_Name = driver.find_element(By.XPATH,'//a[text()="Primary Care Provider"]/following::table[1]//tr[3]/td/dl/dd[1]').text
        except:
            PCP_Name = 'PCP Not Available'
    row[22].value = PCP_Name
    workbook.save(FFM_Bot_Output_File_Path)
    print('Insuarance Type Done')

    try:
        Plan_Description = driver.find_element(By.XPATH,'//a[text()="Active Coverage"]/following::table[1]//tr[td="Health Benefit Plan Coverage"]/td[4]').text
    except:
        try:
            Plan_Description = driver.find_element(By.XPATH,'//a[text()="Active Coverage"]/following::table[1]//tr/th[contains(text(),"Description")]/following::td[4]').text
        except:
            Plan_Description = 'PD Not Available'
    row[38].value = Plan_Description
    workbook.save(FFM_Bot_Output_File_Path)
    print('Plan Description Done')
    
    try:
        coinsurance_val = driver.find_element(By.XPATH,'//a[text()="Co-Insurance"]/following::table[1]//tr[td="Individual" and td="In network" and td ="Professional (Physician) Visit - Office"]/td[5]').text
    except:
        try:
            coinsurance_val = driver.find_element(By.XPATH,'//a[text()="Co-Insurance"]/following::table[1]//tr[./td/span[contains(text(),"OFFICE VISIT - PROFESSIONAL")]]/preceding-sibling::tr[td="Individual" and td="In network"][1]/td[5]').text
        except:
            coinsurance_val = None
    if coinsurance_val != None:
        coinsurance = re.search(pattern,coinsurance_val).group()
        row[31].value = coinsurance
    else:
        row[31].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Co-Insurance Done')

    try:
        copayment_val = driver.find_element(By.XPATH,'//a[text()="Co-Payment"]/following::table[1]//tr[td="Individual" and td="In network" and td ="Professional (Physician) Visit - Office"]/td[5]').text
    except:
        try:
            copayment_val = driver.find_element(By.XPATH,'//a[text()="Co-Payment"]/following::table[1]//tr[./td/span[contains(text(),"OFFICE VISIT - PROFESSIONAL")]]/preceding-sibling::tr[td="Individual" and td="In network"][1]/td[5]').text
        except:
            copayment_val = None
    if copayment_val != None:
        copayment = re.search(pattern,copayment_val).group()
        row[32].value = copayment
    else:
        row[32].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Co-Payment Done')

    try:
        D_individual_per = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr[./td/span[contains(text(),"OFFICE VISIT - PROFESSIONAL")]]/preceding-sibling::tr[td="Individual" and td="In network" and td[contains(text(),"per")]][1]/td[5]').text
    except:
        try:    
            D_individual_per = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr[td="Individual" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"per")][1]').text
        except:
            D_individual_per = None
    if D_individual_per != None:
        d_indiv_per = re.search(pattern,D_individual_per).group()
        row[23].value = d_indiv_per
    else:
        row[23].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Individual Done')

    try:
        D_individual_remaining = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr[./td/span[contains(text(),"OFFICE VISIT - PROFESSIONAL")]]/preceding-sibling::tr[td="Individual" and td="In network" and td[contains(text(),"Remai")]][1]/td[5]').text
    except:
        try:   
            D_individual_remaining = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr[td="Individual" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"Remai")][1]').text
        except:
            D_individual_remaining = None
    if D_individual_remaining != None:
        d_indiv_remain = re.search(pattern,D_individual_remaining).group()
        row[24].value = d_indiv_remain
    else:
        row[24].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Individual Remaining Done')

    try:
        D_family_per = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr[./td/span[contains(text(),"OFFICE VISIT - PROFESSIONAL")]]/preceding-sibling::tr[td="Family" and td="In network" and td[contains(text(),"per")]][1]/td[5]').text
    except:
        try:
            D_family_per = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr[td="Family" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"per")][1]').text
        except:
            D_family_per = None
    if D_family_per != None:
        d_fami_per = re.search(pattern,D_family_per).group()
        row[25].value = d_fami_per
    else:
        row[25].value = 'NA' 
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Family Done')

    try:
        D_family_remaining = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr[./td/span[contains(text(),"OFFICE VISIT - PROFESSIONAL")]]/preceding-sibling::tr[td="Family" and td="In network" and td[contains(text(),"Remai")]][1]/td[5]').text
    except:
        try:
            D_family_remaining = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr[td="Family" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"Remai")][1]').text
        except:
            D_family_remaining = None
    if D_family_remaining != None:
        d_fami_remain = re.search(pattern,D_family_remaining).group()
        row[26].value = d_fami_remain
    else:
        row[26].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Family Remaining Done')

    try:
        OofP_individual_per = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[./td/span[contains(text(),"OFFICE VISIT - PROFESSIONAL")]]/preceding-sibling::tr[td="Individual" and td="In network" and td[contains(text(),"per")]][1]/td[5]').text
    except:
        try:
            OofP_individual_per = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[td="Individual" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"per")][1]').text
        except:
            OofP_individual_per = None
    if OofP_individual_per != None:
        oop_indiv_per = re.search(pattern,OofP_individual_per).group()
        row[27].value = oop_indiv_per
    else:
        row[27].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Out of Pocket Individual Done')

    try:
        OofP_individual_remaining = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[./td/span[contains(text(),"OFFICE VISIT - PROFESSIONAL")]]/preceding-sibling::tr[td="Individual" and td="In network" and td[contains(text(),"Remai")]][1]/td[5]').text
    except:
        try:
            OofP_individual_remaining = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[td="Individual" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"Remai")][1]').text
        except:
            OofP_individual_remaining = None
    if OofP_individual_remaining != None:
        oop_indiv_remain = re.search(pattern,OofP_individual_remaining).group()
        row[28].value = oop_indiv_remain
    else:
        row[28].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Out of Pocket Individual Remaining Done')

    try:
        OofP_family_per = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[./td/span[contains(text(),"OFFICE VISIT - PROFESSIONAL")]]/preceding-sibling::tr[td="Family" and td="In network" and td[contains(text(),"per")]][1]/td[5]').text
    except:
        try:
            OofP_family_per = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[td="Family" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"per")][1]').text
        except:
            OofP_family_per = None
    if OofP_family_per != None:
        oop_fami_per = re.search(pattern,OofP_family_per).group()
        row[29].value = oop_fami_per
    else:
        row[29].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Out of Pocket Family Done')

    try:
        OofP_family_remaining = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[./td/span[contains(text(),"OFFICE VISIT - PROFESSIONAL")]]/preceding-sibling::tr[td="Family" and td="In network" and td[contains(text(),"Remai")]][1]/td[5]').text
    except:
        try:
            OofP_family_remaining = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[td="Family" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"Remai")][1]').text
        except:
            OofP_family_remaining = None

    if OofP_family_remaining != None:
        oop_fami_remain = re.search(pattern,OofP_family_remaining).group()
        row[30].value = oop_fami_remain
    else:
        row[30].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Out of Pocket Family Remaining Done')

def United_Healthcare():
    try:
        plan_Begin_date = driver.find_element(By.XPATH,'//*[@id="BasicProfile"]/dl/dt[text()="Plan Begin Date:"]/following-sibling::dd[1]').text
    except:
        try:
            plan_Begin_date = driver.find_element(By.XPATH,'//*[@id="BasicProfile"]/dl/dt[text()="Eligibility Begin Date:"]/following-sibling::dd').text
        except:
            try:
                plan_Begin_date = driver.find_element(By.XPATH,'//*[@id="BasicProfile"]/dl/dt[contains(text(),"Plan Date")]//following::dd').text
            except:
                plan_Begin_date = 'NA'
    try:
        Plan_end_date = driver.find_element(By.XPATH,'//*[@id="BasicProfile"]/dl/dt[text()="Plan End Date:"]/following-sibling::dd[1]').text
    except:
        Plan_end_date = 'NA'

    if '-' in plan_Begin_date:
        list1 = plan_Begin_date.split('-')
        Plan_Effective_date = list1[0] if len(list1) > 0 else None
        Plan_term_date = list1[1] if len(list1) > 1 else None
        row[18].value = Plan_Effective_date
        row[19].value = Plan_term_date
    else:
        row[18].value = plan_Begin_date
        row[19].value = Plan_end_date
    workbook.save(FFM_Bot_Output_File_Path)
    print('Plan Dates Done')

    InsuranceType = driver.find_element(By.XPATH,'//a[text()="Active Coverage"]/following::table[1]//tr[2]/td[3]').text
    row[21].value = InsuranceType
    if '(HMO)' in InsuranceType:
        row[20].value = 'Yes'
    else:
        row[20].value = 'No'
    try:
        PCP_Name = driver.find_element(By.XPATH,'//a[text()="Primary Care Provider"]/following::table[1]//tr[3]/td/dl/dt[text()="Primary Care Provider"]/following-sibling::dd').text
    except:
        try:
            PCP_Name = driver.find_element(By.XPATH,'//a[text()="Primary Care Provider"]/following::table[1]//tr[3]/td/dl/dd[1]').text
        except:
            PCP_Name = 'PCP Not Available'
    row[22].value = PCP_Name
    workbook.save(FFM_Bot_Output_File_Path)
    print('Insuarance Type Done')

    try:
        Plan_Description = driver.find_element(By.XPATH,'//a[text()="Active Coverage"]/following::table[1]//tr[td="Health Benefit Plan Coverage"]/td[4]').text
    except:
        try:
            Plan_Description = driver.find_element(By.XPATH,'//a[text()="Active Coverage"]/following::table[1]//tr/th[contains(text(),"Description")]/following::td[4]').text
        except:
            Plan_Description = 'PD Not Available'
    row[38].value = Plan_Description
    workbook.save(FFM_Bot_Output_File_Path)
    print('Plan Description Done')
    
    try:
        coinsurance_val = driver.find_element(By.XPATH,'//a[text()="Co-Insurance"]/following::table[1]//tr/td[text()="Health Benefit Plan Coverage"]/following-sibling::td[3]').text
    except:
        try:
            coinsurance_val = driver.find_element(By.XPATH,'//a[text()="Co-Insurance"]/following::table[1]//tr[td/span[contains(text(),"OFFICE VISIT PRIMARY CARE")]]/preceding-sibling::tr[1]/td[5]').text
        except:
            coinsurance_val = None
    if coinsurance_val != None:
        coinsurance = re.search(pattern,coinsurance_val).group()
        row[31].value = coinsurance
    else:
        row[31].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Co-Insurance Done')

    try:
        copayment_val = driver.find_element(By.XPATH,'//a[text()="Co-Payment"]/following::table[1]//tr[td/span[contains(text(),"OFFICE VISIT PRIMARY CARE")]]/preceding-sibling::tr[1]/td[5]').text
    except:
        try:
            copayment_val = driver.find_element(By.XPATH,'//a[text()="Co-Payment"]/following::table[1]//tr[td="Individual" and td[text()="Professional (Physician) Visit - Office"]]/td[5]').text
        except:
            copayment_val = None
    if copayment_val != None:
        copayment = re.search(pattern,copayment_val).group()
        row[32].value = copayment
    else:
        row[32].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Co-Payment Done')

    try:
        D_individual_per = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr/preceding-sibling::tr[td="Individual" and td[contains(text(),"per")]][1]/td[5]').text
    except:
        D_individual_per = None
        
    if D_individual_per != None:
        d_indiv_per = re.search(pattern,D_individual_per).group()
        row[23].value = d_indiv_per
    else:
        row[23].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Individual Done')

    try:
        D_individual_remaining = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr/preceding-sibling::tr[td="Individual" and td[contains(text(),"Remaining")]][1]/td[5]').text
    except:
        D_individual_remaining = None
    if D_individual_remaining != None:
        d_indiv_remain = re.search(pattern,D_individual_remaining).group()
        row[24].value = d_indiv_remain
    else:
        row[24].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Individual Remaining Done')

    try:
        D_family_per = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr/preceding-sibling::tr[td="Family" and td[contains(text(),"per")]][1]/td[5]').text
    except:
        D_family_per = None
    if D_family_per != None:
        d_fami_per = re.search(pattern,D_family_per).group()
        row[25].value = d_fami_per
    else:
        row[25].value = 'NA' 
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Family Done')

    try:
        D_family_remaining = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr[td="Family" and td[contains(text(),"Remai")]][1]/td[5]').text
    except:
        D_family_remaining = None
    if D_family_remaining != None:
        d_fami_remain = re.search(pattern,D_family_remaining).group()
        row[26].value = d_fami_remain
    else:
        row[26].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Family Remaining Done')

    try:
        OofP_individual_per = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[td="Individual" and td[contains(text(),"per")]][1]/td[5]').text
    except:
        OofP_individual_per = None
    if OofP_individual_per != None:
        oop_indiv_per = re.search(pattern,OofP_individual_per).group()
        row[27].value = oop_indiv_per
    else:
        row[27].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Out of Pocket Individual Done')

    try:
        OofP_individual_remaining = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[td="Individual" and td[contains(text(),"Remaining")]][1]/td[5]').text
    except:
        OofP_individual_remaining = None
    if OofP_individual_remaining != None:
        oop_indiv_remain = re.search(pattern,OofP_individual_remaining).group()
        row[28].value = oop_indiv_remain
    else:
        row[28].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Out of Pocket Individual Remaining Done')

    try:
        OofP_family_per = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[td="Family" and td[contains(text(),"per")]][1]/td[5]').text
    except:
        OofP_family_per = None
    if OofP_family_per != None:
        oop_fami_per = re.search(pattern,OofP_family_per).group()
        row[29].value = oop_fami_per
    else:
        row[29].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Out of Pocket Family Done')

    try:
        OofP_family_remaining = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[td="Family" and td[contains(text(),"Remaining")]][1]/td[5]').text
    except:
        OofP_family_remaining = None
    if OofP_family_remaining != None:
        oop_fami_remain = re.search(pattern,OofP_family_remaining).group()
        row[30].value = oop_fami_remain
    else:
        row[30].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Out of Pocket Family Remaining Done')

def Aetna_Data():
    try:
        plan_Begin_date = driver.find_element(By.XPATH,'//*[@id="BasicProfile"]/dl/dt[text()="Plan Begin Date:"]/following-sibling::dd[1]').text
    except:
        try:
            plan_Begin_date = driver.find_element(By.XPATH,'//*[@id="BasicProfile"]/dl/dt[text()="Eligibility Begin Date:"]/following-sibling::dd').text
        except:
            plan_Begin_date = 'NA'
    try:
        Plan_end_date = driver.find_element(By.XPATH,'//*[@id="BasicProfile"]/dl/dt[text()="Plan End Date:"]/following-sibling::dd[1]').text
    except:
        Plan_end_date = 'NA'
    
    if '-' in plan_Begin_date:
        list1 = plan_Begin_date.split('-')
        Plan_Effective_date = list1[0] if len(list1) > 0 else None
        Plan_term_date = list1[1] if len(list1) > 1 else None
        row[18].value = Plan_Effective_date
        row[19].value = Plan_term_date
    else:
        row[18].value = plan_Begin_date
        row[19].value = Plan_end_date
    workbook.save(FFM_Bot_Output_File_Path)
    print('Plan Dates Done')

    InsuranceType = driver.find_element(By.XPATH,'//a[text()="Active Coverage"]/following::table[1]//tr[2]/td[3]').text
    row[21].value = InsuranceType
    if '(HMO)' in InsuranceType:
        row[20].value = 'Yes'
    else:
        row[20].value = 'No'
    try:
        PCP_Name = driver.find_element(By.XPATH,'//a[text()="Primary Care Provider"]/following::table[1]//tr[3]/td/dl/dt[text()="Primary Care Provider"]/following-sibling::dd').text
    except:
        try:
            PCP_Name = driver.find_element(By.XPATH,'//a[text()="Primary Care Provider"]/following::table[1]//tr[3]/td/dl/dd[1]').text
        except:
            PCP_Name = 'PCP Not Available'
    row[22].value = PCP_Name
    workbook.save(FFM_Bot_Output_File_Path)
    print('Insuarance Type Done')

    try:
        Plan_Description = driver.find_element(By.XPATH,'//a[text()="Active Coverage"]/following::table[1]//tr[td="Health Benefit Plan Coverage"]/td[4]').text
    except:
        try:
            Plan_Description = driver.find_element(By.XPATH,'//a[text()="Active Coverage"]/following::table[1]//tr/th[contains(text(),"Description")]/following::td[4]').text
        except:
            Plan_Description = 'PD Not Available'
    row[38].value = Plan_Description
    workbook.save(FFM_Bot_Output_File_Path)
    print('Plan Description Done')

    try:
        coinsurance_val = driver.find_element(By.XPATH,'//a[text()="Co-Insurance"]/following::table[1]//tr[td="Individual" and td="Health Benefit Plan Coverage" and td ="In network"]/td[5]').text
    except:
        try:
            coinsurance_val = driver.find_element(By.XPATH,'//a[text()="Co-Insurance"]/following::table[1]//tr[td="Family" and td="Health Benefit Plan Coverage" and td ="In network"]/td[5]').text
        except:
            try:
                coinsurance_val = driver.find_element(By.XPATH,'//a[text()="Co-Insurance"]/following::table[1]//tr[td="Individual" and td="In network" and td ="Professional (Physician) Visit - Office"]/td[5]').text
            except:
                try:
                    coinsurance_val = driver.find_element(By.XPATH,'//a[text()="Co-Insurance"]/following::table[1]//tr[td="Family" and td="In network" and td ="Professional (Physician) Visit - Office"]/td[5]').text
                except:
                    try:
                        coinsurance_val = driver.find_element(By.XPATH,'//a[text()="Co-Insurance"]/following::table[1]//tr[td="Individual" and td[text()="Professional (Physician) Visit - Office"] and td="In network"]/td[5]').text
                    except:
                        try:
                            coinsurance_val = driver.find_element(By.XPATH,'//a[text()="Co-Insurance"]/following::table[1]//tr[td="Individual" and td[contains(text(),"Professional (Physician)")] and td="In network"]/td[5]').text
                        except:
                            coinsurance_val = None
    if coinsurance_val != None:
        coinsurance = re.search(pattern,coinsurance_val).group()
        row[31].value = coinsurance
    else:
        row[31].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Co-Insurance Done')

    try:
        copayment_val = driver.find_element(By.XPATH,'//a[text()="Co-Payment"]/following::table[1]//tr[td="Individual" and td="In network" and td ="Professional (Physician) Visit - Office"]/td[5]').text
    except:
        try:
            copayment_val = driver.find_element(By.XPATH,'//a[text()="Co-Payment"]/following::table[1]//tr[td="Family" and td="In network" and td ="Professional (Physician) Visit - Office"]/td[5]').text
        except:
            try:
                copayment_val = driver.find_element(By.XPATH,'//a[text()="Co-Payment"]/following::table[1]//tr[td="Individual" and td="In network" and td[text()="Professional (Physician) Visit - Office"]]/td[5]').text
            except:
                try:
                    copayment_val = driver.find_element(By.XPATH,'//a[text()="Co-Payment"]/following::table[1]//tr[td="Individual" and td[contains(text(),"Professional (Physician)")] and td="In network"]/td[5]').text
                except:
                    copayment_val = None
    if copayment_val != None:
        copayment = re.search(pattern,copayment_val).group()
        row[32].value = copayment
    else:
        row[32].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Co-Payment Done')

    try:    
        D_individual_per = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr[td="Individual" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"per")][1]').text
    except:
        D_individual_per = None
    if D_individual_per != None:
        d_indiv_per = re.search(pattern,D_individual_per).group()
        row[23].value = d_indiv_per
    else:
        row[23].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Individual Done')
    
    try:   
        D_individual_remaining = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr[td="Individual" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"Remai")][1]').text
    except:
        D_individual_remaining = None
    if D_individual_remaining != None:
        d_indiv_remain = re.search(pattern,D_individual_remaining).group()
        row[24].value = d_indiv_remain
    else:
        row[24].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Individual Remaining Done')

    try:
        D_family_per = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr[td="Family" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"per")][1]').text
    except:
        D_family_per = None
    if D_family_per != None:
        d_fami_per = re.search(pattern,D_family_per).group()
        row[25].value = d_fami_per
    else:
        row[25].value = 'NA' 
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Family Done')
    
    try:
        D_family_remaining = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr[td="Family" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"Remai")][1]').text
    except:
        D_family_remaining = None
    if D_family_remaining != None:
        d_fami_remain = re.search(pattern,D_family_remaining).group()
        row[26].value = d_fami_remain
    else:
        row[26].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Family Remaining Done')
    
    try:
        OofP_individual_per = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[td[text()="Individual"] and td[text()="In network"] and td[text()="Health Benefit Plan Coverage"] and not(td[contains(text(),"Remaining")])]/td[5]').text
    except:
        OofP_individual_per = None
    if OofP_individual_per != None:
        oop_indiv_per = re.search(pattern,OofP_individual_per).group()
        row[27].value = oop_indiv_per
    else:
        row[27].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Out of Pocket Individual Done')

    try:
        OofP_individual_remaining = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[td="Individual" and td="In network" and td ="Health Benefit Plan Coverage" and td[contains(text(),"Remaining")]]/td[5]').text
    except:
        OofP_individual_remaining = None
    if OofP_individual_remaining != None:
        oop_indiv_remain = re.search(pattern,OofP_individual_remaining).group()
        row[28].value = oop_indiv_remain
    else:
        row[28].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Out of Pocket Individual Remaining Done')

    try:
        OofP_family_per = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[td[text()="Family"] and td[text()="In network"] and td[text()="Health Benefit Plan Coverage"] and not(td[contains(text(),"Remaining")])]/td[5]').text
    except:
        OofP_family_per = None
    if OofP_family_per != None:
        oop_fami_per = re.search(pattern,OofP_family_per).group()
        row[29].value = oop_fami_per
    else:
        row[29].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Out of Pocket Family Done')

    try:
        OofP_family_remaining = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[td="Family" and td="In network" and td ="Health Benefit Plan Coverage" and td[contains(text(),"Remaining")]]/td[5]').text
    except:
        OofP_family_remaining = None
    if OofP_family_remaining != None:
        oop_fami_remain = re.search(pattern,OofP_family_remaining).group()
        row[30].value = oop_fami_remain
    else:
        row[30].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Out of Pocket Family Remaining Done')

def MEDICARE():
    try:
        Policy_Effective_date = driver.find_element(By.XPATH,'//*[@id="panel_benefitinformation"]/a[text()="Active Coverage"]/following::table[1]//tr[td[text()="Medicare Part B"]]/following-sibling::tr/td/dl/dd').text
    except:
        Policy_Effective_date = 'NA'
    row[18].value = Policy_Effective_date
    row[19].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Plan Dates Done')

    try:
        InsuranceType = driver.find_element(By.XPATH,'//*[@id="panel_benefitinformation"]/a[text()="Contact Following Entity for Eligibility or Benefit Information"]/following::table//tr/td[3]').text
    except:
        InsuranceType = driver.find_element(By.XPATH,'//*[@id="panel_benefitinformation"]/a[text()="Co-Insurance"]/following::table[1]//tr[td="Health Benefit Plan Coverage"]/td[3]').text
    row[21].value = InsuranceType
    if '(HMO)' in InsuranceType or '(PPO)' in InsuranceType:
        Primary_Payer = driver.find_element(By.XPATH,'//a[text()="Contact Following Entity for Eligibility or Benefit Information"]/following::table//dt[text()="Primary Payer"]/following-sibling::dd').text
        Message = driver.find_element(By.XPATH,'//a[text()="Contact Following Entity for Eligibility or Benefit Information"]/following::table//tr/td/span').text
        row[35].value = 'Yes'
        if '(HMO)' in InsuranceType:
            row[20].value = 'Yes'
        else:
            row[20].value = 'No'
        row[36].value = Primary_Payer
        row[37].value = Message
    else:
        row[35].value = 'No'
        row[36].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Insuarance Type Done')

    try:
        coinsurance_val = driver.find_element(By.XPATH,'//*[@id="panel_benefitinformation"]/a[text()="Co-Insurance"]/following::table[1]//tr[td="Health Benefit Plan Coverage" and td="Medicare Part B"]/td[5]').text
    except:
        coinsurance_val = None
    if coinsurance_val != None:
        coinsurance = re.search(pattern,coinsurance_val).group()
        row[31].value = coinsurance
    else:
        row[31].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Co-Insurance Done')

    try:
        D_individual_per = driver.find_element(By.XPATH,'//*[@id="panel_benefitinformation"]/a[text()="Deductible"]/following::table[1]//tr[td="Health Benefit Plan Coverage" and td="Medicare Part B" and td[contains(text(),"per")]]/td[5]').text
    except:
        D_individual_per = None
    if D_individual_per != None:
        d_indiv_per = re.search(pattern,D_individual_per).group()
        row[23].value = d_indiv_per
    else:
        row[23].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Individual Done')

    try:
        D_individual_remaining = driver.find_element(By.XPATH,'//*[@id="panel_benefitinformation"]/a[text()="Deductible"]/following::table[1]//tr[td="Health Benefit Plan Coverage" and td="Medicare Part B" and td[contains(text(),"Remaining")]]/td[5]').text
    except:
        D_individual_remaining = None
    if D_individual_remaining != None:
        d_indiv_remain = re.search(pattern,D_individual_remaining).group()
        row[24].value = d_indiv_remain
    else:
        row[24].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Individual Remaining Done')


def Common_Data():
    try:
        plan_Begin_date = driver.find_element(By.XPATH,'//*[@id="BasicProfile"]/dl/dt[text()="Plan Begin Date:"]/following-sibling::dd[1]').text
    except:
        try:
            plan_Begin_date = driver.find_element(By.XPATH,'//*[@id="BasicProfile"]/dl/dt[text()="Eligibility Begin Date:"]/following-sibling::dd').text
        except:
            plan_Begin_date = 'NA'
    try:
        Plan_end_date = driver.find_element(By.XPATH,'//*[@id="BasicProfile"]/dl/dt[text()="Plan End Date:"]/following-sibling::dd[1]').text
    except:
        Plan_end_date = 'NA'
    
    if '-' in plan_Begin_date:
        list1 = plan_Begin_date.split('-')
        Plan_Effective_date = list1[0] if len(list1) > 0 else None
        Plan_term_date = list1[1] if len(list1) > 1 else None
        row[18].value = Plan_Effective_date
        row[19].value = Plan_term_date
    else:
        row[18].value = plan_Begin_date
        row[19].value = Plan_end_date
    workbook.save(FFM_Bot_Output_File_Path)
    print('Plan Dates Done')

    InsuranceType = driver.find_element(By.XPATH,'//a[text()="Active Coverage"]/following::table[1]//tr[2]/td[3]').text
    row[21].value = InsuranceType
    if '(HMO)' in InsuranceType:
        row[20].value = 'Yes'
    else:
        row[20].value = 'No'
    try:
        PCP_Name = driver.find_element(By.XPATH,'//a[text()="Primary Care Provider"]/following::table[1]//tr[3]/td/dl/dt[text()="Primary Care Provider"]/following-sibling::dd').text
    except:
        try:
            PCP_Name = driver.find_element(By.XPATH,'//a[text()="Primary Care Provider"]/following::table[1]//tr[3]/td/dl/dd[1]').text
        except:
            PCP_Name = 'PCP Not Available'
    row[22].value = PCP_Name
    workbook.save(FFM_Bot_Output_File_Path)
    print('Insuarance Type Done')

    try:
        Plan_Description = driver.find_element(By.XPATH,'//a[text()="Active Coverage"]/following::table[1]//tr[td="Health Benefit Plan Coverage"]/td[4]').text
    except:
        try:
            Plan_Description = driver.find_element(By.XPATH,'//a[text()="Active Coverage"]/following::table[1]//tr/th[contains(text(),"Description")]/following::td[4]').text
        except:
            Plan_Description = 'PD Not Available'
    row[38].value = Plan_Description
    workbook.save(FFM_Bot_Output_File_Path)
    print('Plan Description Done')

    try:
        coinsurance_val = driver.find_element(By.XPATH,'//a[text()="Co-Insurance"]/following::table[1]//tr[td="Individual" and td="Health Benefit Plan Coverage" and td ="In network"]/td[5]').text
    except:
        try:
            coinsurance_val = driver.find_element(By.XPATH,'//a[text()="Co-Insurance"]/following::table[1]//tr[td="Family" and td="Health Benefit Plan Coverage" and td ="In network"]/td[5]').text
        except:
            try:
                coinsurance_val = driver.find_element(By.XPATH,'//a[text()="Co-Insurance"]/following::table[1]//tr[td="Individual" and td="In network" and td ="Professional (Physician) Visit - Office"]/td[5]').text
            except:
                try:
                    coinsurance_val = driver.find_element(By.XPATH,'//a[text()="Co-Insurance"]/following::table[1]//tr[td="Family" and td="In network" and td ="Professional (Physician) Visit - Office"]/td[5]').text
                except:
                    try:
                        coinsurance_val = driver.find_element(By.XPATH,'//a[text()="Co-Insurance"]/following::table[1]//tr[td="Individual" and td[text()="Professional (Physician) Visit - Office"] and td="In network"]/td[5]').text
                    except:
                        try:
                            coinsurance_val = driver.find_element(By.XPATH,'//a[text()="Co-Insurance"]/following::table[1]//tr[td="Individual" and td[contains(text(),"Professional (Physician)")] and td="In network"]/td[5]').text
                        except:
                            coinsurance_val = None
    if coinsurance_val != None:
        coinsurance = re.search(pattern,coinsurance_val).group()
        row[31].value = coinsurance
    else:
        row[31].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Co-Insurance Done')

    try:
        copayment_val = driver.find_element(By.XPATH,'//a[text()="Co-Payment"]/following::table[1]//tr[td="Individual" and td="In network" and td ="Professional (Physician) Visit - Office"]/td[5]').text
    except:
        try:
            copayment_val = driver.find_element(By.XPATH,'//a[text()="Co-Payment"]/following::table[1]//tr[td="Family" and td="In network" and td ="Professional (Physician) Visit - Office"]/td[5]').text
        except:
            try:
                copayment_val = driver.find_element(By.XPATH,'//a[text()="Co-Payment"]/following::table[1]//tr[td="Individual" and td="In network" and td[text()="Professional (Physician) Visit - Office"]]/td[5]').text
            except:
                try:
                    copayment_val = driver.find_element(By.XPATH,'//a[text()="Co-Payment"]/following::table[1]//tr[td="Individual" and td[contains(text(),"Professional (Physician)")] and td="In network"]/td[5]').text
                except:
                    copayment_val = None
    if copayment_val != None:
        copayment = re.search(pattern,copayment_val).group()
        row[32].value = copayment
    else:
        row[32].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Co-Payment Done')

    try:    
        D_individual_per = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr[td="Individual" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"per")][1]').text
    except:
        D_individual_per = None
    if D_individual_per != None:
        d_indiv_per = re.search(pattern,D_individual_per).group()
        row[23].value = d_indiv_per
    else:
        row[23].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Individual Done')
    
    try:   
        D_individual_remaining = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr[td="Individual" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"Remai")][1]').text
    except:
        D_individual_remaining = None
    if D_individual_remaining != None:
        d_indiv_remain = re.search(pattern,D_individual_remaining).group()
        row[24].value = d_indiv_remain
    else:
        row[24].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Individual Remaining Done')

    try:
        D_family_per = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr[td="Family" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"per")][1]').text
    except:
        D_family_per = None
    if D_family_per != None:
        d_fami_per = re.search(pattern,D_family_per).group()
        row[25].value = d_fami_per
    else:
        row[25].value = 'NA' 
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Family Done')
    
    try:
        D_family_remaining = driver.find_element(By.XPATH,'//a[text()="Deductible"]/following::table[1]//tr[td="Family" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"Remai")][1]').text
    except:
        D_family_remaining = None
    if D_family_remaining != None:
        d_fami_remain = re.search(pattern,D_family_remaining).group()
        row[26].value = d_fami_remain
    else:
        row[26].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Deductible Family Remaining Done')
    
    try:
        OofP_individual_per = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[td="Individual" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"per")][1]').text
    except:
        OofP_individual_per = None
    if OofP_individual_per != None:
        oop_indiv_per = re.search(pattern,OofP_individual_per).group()
        row[27].value = oop_indiv_per
    else:
        row[27].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Out of Pocket Individual Done')

    try:
        OofP_individual_remaining = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[td="Individual" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"Remai")][1]').text
    except:
        OofP_individual_remaining = None
    if OofP_individual_remaining != None:
        oop_indiv_remain = re.search(pattern,OofP_individual_remaining).group()
        row[28].value = oop_indiv_remain
    else:
        row[28].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Out of Pocket Individual Remaining Done')

    try:
        OofP_family_per = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[td="Family" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"per")][1]').text
    except:
        OofP_family_per = None
    if OofP_family_per != None:
        oop_fami_per = re.search(pattern,OofP_family_per).group()
        row[29].value = oop_fami_per
    else:
        row[29].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Out of Pocket Family Done')

    try:
        OofP_family_remaining = driver.find_element(By.XPATH,'//a[text()="Out of Pocket (Stop Loss)"]/following::table[1]//tr[td="Family" and td="In network" and td ="Health Benefit Plan Coverage"]/td[contains(text(),"Remai")][1]').text
    except:
        OofP_family_remaining = None
    if OofP_family_remaining != None:
        oop_fami_remain = re.search(pattern,OofP_family_remaining).group()
        row[30].value = oop_fami_remain
    else:
        row[30].value = 'NA'
    workbook.save(FFM_Bot_Output_File_Path)
    print('Out of Pocket Family Remaining Done')

def Description_PlanType():
    description_and_type = {
        "000003H STAR": "STAR",
        "27248TX001000100 COMMUNITY PREMIER GOLD MARKETPLAC": "HMO",
        "27248TX001000300 COMMUNITY PREMIER BRONZE MARKETPL": "HMO",
        "27248TX001000500 COMMUNITY PREMIER GOLD MARKETPLAC": "HMO",
        "27248TX001001800 COMMUNITY PREMIER BRONZE MARKETPL": "HMO",
        "ABCBS REGULAR BUSINESS PPO GROUP": "PPO",
        "ADVANTAGE SALES AND MARKETING EPO LLC ASO BLUE SH": "EPO",
        "Aetna Choice POS II": "Choice POS II",
        "Aetna Medicare Choice II Plan (PPO)": "MDC PPO",
        "Aetna Medicare Choice Plan (PPO)": "MDC PPO",
        "Aetna Medicare Freedom Plan (PPO)": "MDC PPO",
        "Aetna Medicare Prime Plan (HMO)": "MDC ADV HMO",
        "Aetna Select": "EPO",
        "AHF Choice POS II": "Choice POS II",
        "ALLIANCE SELECT": "PPO",
        "BASIC": "PPO",
        "BASIC SECONDARY": "PPO",
        "BC PPO": "PPO",
        "BC PPO EXCLUSIVE": "PPO",
        "BC PPO HSA COMPATIBLE": "PPO",
        "BC PPO INCENTIVE": "PPO",
        "BCBS PPO HDHP BD": "PPO",
        "BCBS PPO HDHP MED/SURG": "PPO",
        "BLUE ACCESS PPO": "PPO",
        "BLUE ACCESS PPO BCBS UNITED WISCONSIN": "PPO",
        "Blue Advantage PPO": "PPO",
        "BLUE CARD INCENTIVE NO SUTTER": "PPO",
        "BLUE CARD PPO 80/50- MED SURG": "PPO",
        "BLUE CARD PPO HDHP BD": "PPO",
        "BLUE CARD PPO HDHP MED/SURG": "PPO",
        "BLUE CHOICE ADVANTAGE": "PPO",
        "Blue Options": "PPO",
        "Blue Saver": "PPO",
        "BLUECARD PPO": "PPO",
        "BLUECHOICE MEDICAL": "PPO",
        "BlueChoice PPO": "PPO",
        "BLUEOPTIONS EVERYDAY HEALTH 19103": "PPO",
        "BlueOptions Everyday Health 22101": "PPO",
        "BLUEOPTIONS MEDICAL": "PPO",
        "BLUEOPTIONS PREDICTABLE COST 03559": "PPO",
        "BLUEPREFERRED": "PPO",
        "CA GNRIC ANTHEM PPO ELEMENTS CHOICE HSA": "PPO",
        "CA SG ANTHEM HSA": "PPO",
        "CDHP CA GENERIC HLTH SAVINGS ACCT": "PPO",
        "CDHP CT BC PPO HEALTH SAVINGS ACCOUNT": "PPO",
        "CDHP GEORGIA HEALTH SAVINGS ACCOUNT": "PPO",
        "CDHP HSA PLUS PPO INDIANA": "PPO",
        "CDHP IN HEALTH SAVINGS ACCOUNT": "PPO",
        "CDHP NY NAT GENC HEALTH SAVINGS ACCT (HSA)": "PPO",
        "CDHP OHIO HEALTH SAVINGS ACCOUNT": "PPO",
        "CDHP VIRGINIA HEALTH SAVINGS ACCOUNT": "PPO",
        "CHOICE": "Choice",
        "Choice": "Choice",
        "Choice Fund HRA Open Access Plus": "Open Access Plus",
        "Choice Fund HSA Open Access Plus": "Open Access Plus",
        "Choice Fund HSA Open Access Plus - In Network": "Open Access Plus",
        "CHOICE PLUS": "Choice Plus",
        "Cigna Alliance Medicare (HMO)": "HMO",
        "Cigna Courage Medicare (HMO)": "HMO",
        "Cigna Preferred Medicare (HMO)": "HMO",
        "Cigna Preferred Savings Medicare (HMO)": "HMO",
        "Cigna TotalCare (HMO D-SNP)": "HMO",
        "Cigna True Choice Medicare (PPO)": "MDC PPO",
        "Cigna True Choice Plus Medicare (PPO)": "MDC PPO",
        "COLORADO BLUE CLASSIC PPO": "PPO",
        "CT BLUE CARD PREFERRED PPO": "PPO",
        "Dental PPO": "PPO",
        "DEVOTED CORE GREATER HOUSTON (HMO)": "HMO",
        "DEVOTED GIVEBACK GREATER HOUSTON (HMO)": "HMO",
        "DEVOTED PRIME GREATER HOUSTON (HMO)": "HMO",
        "Dual Liberty (HMO D-SNP) (H0174006000)": "HMO",
        "EMPIRE BRONZE EPO 20/50 6100 50% W/HSA": "EPO",
        "EMPIRE PPO WITH HSA": "PPO",
        "EMPIRE TOTAL BLUE HSA": "PPO",
        "EXCLUSIVE PROVIDER ORGANIZATION": "EPO",
        "EXCLUSIVE PROVIDER ORGANIZATION MEDICAL": "EPO",
        "GA POS BLUECHOICE OA": "POS",
        "GEN26/COPAY SEL": "PPO",
        "GEORGIA HEALTH SAVINGS ACCOUNT": "PPO",
        "Giveback (HMO) (H0174019000)": "HMO",
        "GOLDEN RULE DENTAL PREMIER ELITE (YEAR 3) PLAN TS3": "HMO",
        "HEALTH MAINTENANCE ORGANIZATION MEDICAL": "HMO",
        "HMOPOS-AARP MEDICARE ADVANTAGE PLAN 1 (HMO-POS)": "HMO-POS",
        "HMOPOS-AARP MEDICARE ADVANTAGE WALGREENS PLAN 1 (H": "HMO-POS",
        "HPHC MA HDHP PPO 0% OP 100% IP AFTER DED PY FAM": "PPO",
        "HRA": "Open Access Plus",
        "HSA Aetna Choice POS II": "Choice POS II",
        "HSA Open Access MC": "Open Access Plus",
        "HSA Qualified HDHP": "Open Access Plus",
        "Humana Gold Plus": "HMO",
        "Humana Honor": "MDC PPO",
        "HUMANA INC NPOS HDHP3": "POS",
        "HumanaChoice": "MDC PPO",
        "INDEMNITY": "MDC PPO",
        "LIS HDHP": "PPO",
        "LocalPlus": "Local Plus",
        "LPPO-AARP MEDICARE ADVANTAGE CHOICE (PPO)": "MDC ADV PPO",
        "LPPO-UNITEDHEALTHCARE GROUP MEDICARE ADVANTAGE (PP": "MDC ADV PPO",
        "LUMENOS HEALTH SAVINGS ACCOUNT": "PPO",
        "MARYLAND PPO": "PPO",
        "MED ADV": "PPO",
        "MED/SURG - PPO": "PPO",
        "Medicare (C04) ESA PPO": "MDC PPO",
        "Medicare (C05) ESA PPO": "MDC PPO",
        "Medicare (P01) ESA PPO": "MDC PPO",
        "Medicare (P01) PPO": "MDC PPO",
        "Medicare (S01) ESA PPO": "MDC PPO",
        "Medicare (V02) ESA PPO": "MDC PPO",
        "Medicare (V03) PPO": "MDC PPO",
        "MEDICARE BLUE PPO LARGE GROUP OPTION 4 (PPO)": "MDC PPO",
        "NB, Health Care Information Systems 2023, BlueFlex": "PPO",
        "No Premium (HMO) (H0174010000)": "HMO",
        "OAP": "OAP",
        "Open Access Aetna Select": "EPO",
        "Open Access Elect Choice": "EPO",
        "Open Access MC": "POS",
        "Open Access Plus": "Open Access Plus",
        "Open Access Plus - In Network": "Open Access Plus",
        "Open Access Plus HDHP": "Open Access Plus",
        "Open Choice": "PPO",
        "Open Choice PPO": "PPO",
        "OPTIONS PPO": "PPO",
        "PERSONAL CHOICE PPO FLEX MED/SURG 2001": "PPO",
        "POINT OF SERVICE MEDICAL": "POS",
        "POINT OF SERVICE OPEN ACCESS MEDICAL": "POS",
        "PPO": "PPO",
        "PPO - ADVANTAGE BLUE": "PPO",
        "PPO - BLUE CARE ELECT DEDUCTIBLE WITH COINSURANCE": "PPO",
        "PPO - BLUE CARE ELECT ENHANCED VALUE OPTION": "PPO",
        "PPO - BLUE CARE ELECT PREFERRED": "PPO",
        "PPO - BLUE CARE ELECT PREFERRED 80 WITH COPAY": "PPO",
        "PPO - BLUE CARE ELECT SAVER WITH COINSURANCE": "PPO",
        "PPO BLUE MED/SURG": "PPO",
        "PPO DENTAL": "PPO",
        "PPO Dental 2000": "PPO",
        "PPO GROUP": "PPO",
        "PPO Plan, Rx1": "PPO",
        "PPO PRUDENT BUYER CLASSIC": "PPO",
        "PPO PRUDENT BUYER INCENTIVE": "PPO",
        "PREFERRED CARE": "PPO",
        "PREFERRED CARE BLUE PBM8214A AT00": "PPO",
        "PREFERRED CARE BLUE PBMHE87A AV54": "PPO",
        "PREFERRED PROVIDER OPTION MEDICAL": "PPO",
        "PREFERRED PROVIDER OPTION PLUS MEDICAL": "PPO",
        "Preferred Provider Organization": "PPO",
        "PREFERRED PROVIDER ORGANIZATION (PPO)": "PPO",
        "Preferred Provider Organization HSA": "PPO",
        "Preferred Provider Plan": "PPO",
        "PREFERRED_PROVIDER_ORG_(PPO)": "PPO",
        "RPPO-UNITEDHEALTHCARE MEDICARE ADVANTAGE CHOICE (R": "MDC ADV PPO",
        "RPPO-UNITEDHEALTHCARE MEDICARE GOLD (REGIONAL PPO": "MDC ADV PPO",
        "RYDER SCOTT COMPANY LP": "PPO",
        "SEHBP Educators Medicare Plan": "Medicare PPO",
        "Select PPO": "PPO",
        "SELF-FUNDED EXP SHARED DED PLAN": "PPO",
        "SELF-FUNDED EXP STANDARD PLAN": "PPO",
        "SELF-FUNDED HDHP HEALTH SAVINGS PLAN": "PPO",
        "SELF-FUNDED HDHP HERITAGE": "PPO",
        "SELF-FUNDED PPO HERITAGE": "PPO",
        "Simply Blue PPO": "PPO",
        "SOLUTION PPO": "PPO",
        "SOURCE CODE HOLDINGS PPO PLAN": "PPO",
        "Specialty No Premium (HMO C-SNP) (H0174008000)": "PPO",
        "STANDARD": "PPO",
        "STANDARD SECONDARY": "PPO",
        "TexanPlus Classic No Premium (HMO) (H4506003000)": "HMO",
        "THALES USA, INC.": "PPO",
        "THE BROOKWOOD COMMUNITY INC": "PPO",
        "THE WONDERFUL COMPANY SHARED ADVANTAGE NATIONAL PP": "PPO",
        "TOTAL FINANCE USA, INC.": "Choice POS",
        "TRADER JOE S CUSTOM PPO 25 500": "PPO",
        "TRICARE SELECT RETIREE": "Tricare Select",
        "TRUE BLUE PPO - HSA QUALIFIED PLAN": "PPO",
        "TX CR NPOS 21 SEP ACC&CPY OV,": "POS",
        "TX H2593-032-000 AMERIVANTAGE DUAL COORDINATION (H": "HMO",
        "TX H8343-001-000 AMERIVANTAGE CHOICE (PPO)": "PPO",
        "TX H8849-008-001 AMERIVANTAGE CLASSIC PLUS (HMO)": "HMO",
        "TX H8849-010-001 AMERIVANTAGE DUAL COORDINATION PL": "HMO",
        "TX HUMANA NPOS LFP 17 RX4": "Open Acces POS",
        "TX HUMANA NPOS LFP 19 COPAY RX": "POS",
        "TX LG NPOS 14-SEP ACC&CPY OV&D": "POS",
        "TX NCR NPOS 16-SEP ACC&CPY OV&": "POS",
        "TX UNITEDHEALTHCARE DUAL COMPLETE HMOPOS H4514-013": "HMO-POS",
        "UNITEDHEALTHCARE CHARTER": "PPO",
        "UNITEDHEALTHCARE CHOICE PLUS": "Choice Plus",
        "UNITEDHEALTHCARE NEXUSACO OA": "Nexus OA PPO",
        "Value Script (PDP) (S4802155)": "PPO",
        "WALMART - HRA - BC/BS BLUECARD PPO": "PPO",
        "West Adv Plan 125": "MDC PPO"
        }

    description = row[38].value

    if description in description_and_type:
        plan_type = description_and_type[description]
    else:
        plan_type = None
    return plan_type


def Commercial_Notes():

    if row[23].value == 'NA':
        Ind_Ded = "Individual Deductible - NA\n"
    else:
        Ind_Ded = f"Individual Deductible - {row[23].value} ({row[24].value} Remaining)\n"

    if row[25].value == 'NA':
        Fam_Ded = "Family Deductible - NA\n"
    else:
        Fam_Ded = f"Family Deductible - {row[25].value} ({row[26].value} Remaining)\n"

    if row[27].value == 'NA':
        Ind_OOP = "Individual OOP - NA\n"
    else:
        Ind_OOP = f"Individual OOP - {row[27].value} ({row[28].value} Remaining)\n"
    
    if row[29].value == 'NA':
        Fam_OOP = "Family OOP - NA\n"
    else:
        Fam_OOP = f"Family OOP - {row[29].value} ({row[30].value} Remaining)\n"

    result = Description_PlanType()

    if row[31].value != None:
        coin = 'Y'
    else:
        coin = 'N'
    if row[32].value != None:
        copay = 'Y'
    else:
        copay = 'N'

    if ((row[23].value == '$0' or row[23].value == 'NA') and (row[24].value == '$0' or row[24].value == 'NA') and (row[25].value == '$0' or row[25].value == 'NA') and (row[26].value == '$0' or row[26].value == 'NA') and (row[27].value == '$0' or row[27].value == 'NA') and (row[28].value == '$0' or row[28].value == 'NA') and (row[29].value == '$0' or row[29].value == 'NA') and (row[30].value == '$0' or row[30].value == 'NA') and (row[31].value == '0%' or row[31].value == 'NA') and (row[32].value == '$0' or row[32].value == 'NA')):
        line_here = 'PLEASE DO NOT COLLECT ANYTHING FROM THE PATIENT'
    elif ((row[28].value == '$0' or row[28].value == '$0.00' or row[28].value == 'NA') or (row[30].value == '$0' or row[30].value == '$0.00' or row[30].value == 'NA')):
        line_here = 'PLEASE DO NOT COLLECT ANYTHING FROM THE PATIENT'
    elif ((row[31].value == '0%' or row[31].value == 'NA') and (row[32].value == '$0' or row[32].value == 'NA') and (row[24].value == '$0' or row[24].value == 'NA' or row[24].value == '$0.00') or (row[26].value == '$0' or row[26].value == '$0.00' or row[26].value == 'NA')):
        line_here = 'PLEASE DO NOT COLLECT ANYTHING FROM THE PATIENT'
    else:
        if row[32].value != None and row[32].value != 'NA':
            line_here = f"PLEASE COLLECT - {row[32].value} COPAYMENT"
        else:
            if row[24].value == '$0' or row[26].value == '$0':
                line_here = f"PLEASE COLLECT - {row[31].value} COINSURANCE"
            else:
                line_here = f"PLEASE COLLECT - DEDUCTIBLE"

    if row[9].value == 'AWV' or row[9].value == 'HRA':
        row[33].value = f"DOS - {Appoinment_date}\n"\
            f"{row[13].value}\n"\
            f"Effective Date - {row[18].value}\n"\
            f"Plan Type - {result}\n"\
            f"Patient is eligible for AWV (CPT-G0438,G0439)covered at 100%? - (Y/N)\n"\
            f"Last AWV DOS: - \n"\
            f"Verified Through TriZetto on {today_date}"
        print('AWV_and_HRA Notes Updated')

    elif row[9].value == 'NP PX AD' or row[9].value == 'Px Adult':
        if row[20].value == 'Yes':
            row[33].value = f"DOS - {Appoinment_date}\n"\
                f"{row[13].value}\n"\
                f"Effective Date - {row[18].value}\n"\
                f"Plan Type - {result}\n"\
                f"PCP Required - {row[20].value}\n"\
                f"Referral Required - {row[20].value}\n"\
                f"PCP: {row[22].value}\n"\
                f"Patient eligible for AD-PX? - (Y/N)\n"\
                f"Covered at 100%\n"\
                f"Last AD-PX date: - \n"\
                f"Screen Labs with Preventive ICD Z00.00 same day with AD-PX? - (Y/N)\n"\
                f"EKG services with Preventive ICD Z00.00 same day with AD-PX? - (Y/N)\n"\
                f"Spoke with - \n"\
                f"Ref# - \n"\
                f"Verified Through TriZetto on {today_date}"
            print('NP_PX_AD_and_Px_Adult Notes Updated')
        else:
            row[33].value = f"DOS - {Appoinment_date}\n"\
                f"{row[13].value}\n"\
                f"Effective Date - {row[18].value}\n"\
                f"Plan Type - {result}\n"\
                f"PCP Required - {row[20].value}\n"\
                f"Referral Required - {row[20].value}\n"\
                f"Patient eligible for AD-PX? - (Y/N)\n"\
                f"Covered at 100%\n"\
                f"Last AD-PX date: - \n"\
                f"Screen Labs with Preventive ICD Z00.00 same day with AD-PX? - (Y/N)\n"\
                f"EKG services with Preventive ICD Z00.00 same day with AD-PX? - (Y/N)\n"\
                f"Spoke with - \n"\
                f"Ref# - \n"\
                f"Verified Through TriZetto on {today_date}"
            print('NP_PX_AD_and_Px_Adult Notes Updated')

    elif row[9].value == 'NP WCC' or row[9].value == 'WCC':
        if row[20].value == 'Yes':
            row[33].value = f"DOS - {Appoinment_date}\n"\
                f"{row[13].value}\n"\
                f"Effective Date - {row[18].value}\n"\
                f"Plan Type - {result}\n"\
                f"PCP Required - {row[20].value}\n"\
                f"Referral Required - {row[20].value}\n"\
                f"PCP: {row[22].value}\n"\
                f"Patient eligible for WCC? - (Y/N)\n"\
                f"Covered at 100%\n"\
                f"Visits allowed as per patients age? 1 VISIT\n"\
                f"Last WCC date:-\n"\
                f"Immunizations codes are covered at 100%.\n"\
                f"Copay Applicable? - ({copay})\n"\
                f"Deductible Applicable? - (Y)\n"\
                f"Co-ins Applicable? - ({coin})\n"\
                f"CPT- 99172 - \n"\
                f"CPT- 96110 - \n"\
                f"Spoke with - \n"\
                f"Ref# - \n"\
                f"Verified Through TriZetto on {today_date}"
            print('NP_WCC_and_WCC Notes Updated')
        else:
            row[33].value = f"DOS - {Appoinment_date}\n"\
                f"{row[13].value}\n"\
                f"Effective Date - {row[18].value}\n"\
                f"Plan Type - {result}\n"\
                f"PCP Required - {row[20].value}\n"\
                f"Referral Required - {row[20].value}\n"\
                f"Patient eligible for WCC? - (Y/N)\n"\
                f"Covered at 100%\n"\
                f"Visits allowed as per patients age? 1 VISIT\n"\
                f"Last WCC date:-\n"\
                f"Immunizations codes are covered at 100%.\n"\
                f"Copay Applicable? - ({copay})\n"\
                f"Deductible Applicable? - (Y)\n"\
                f"Co-ins Applicable? - ({coin})\n"\
                f"CPT- 99172 - \n"\
                f"CPT- 96110 - \n"\
                f"Spoke with - \n"\
                f"Ref# - \n"\
                f"Verified Through TriZetto on {today_date}"
            print('NP_WCC_and_WCC Notes Updated')


    elif row[9].value == 'WWE' or row[9].value == 'NP WWE':
        if row[20].value == 'Yes':
            row[33].value = f"DOS - {Appoinment_date}\n"\
                f"{row[13].value}\n"\
                f"Effective Date - {row[18].value}\n"\
                f"Plan Type - {result}\n"\
                f"PCP Required - {row[20].value}\n"\
                f"Referral Required - {row[20].value}\n"\
                f"PCP: {row[22].value}\n"\
                f"Patient eligible for WWE? - (Y/N)\n"\
                f"Covered at 100%\n"\
                f"Last WWE visit:-\n"\
                f"Spoke with - \n"\
                f"Ref# - \n"\
                f"Verified Through TriZetto on {today_date}"
            print('WWE_and_NP_WWE Notes Updated')
        else:
            row[33].value = f"DOS - {Appoinment_date}\n"\
                f"{row[13].value}\n"\
                f"Effective Date - {row[18].value}\n"\
                f"Plan Type - {result}\n"\
                f"PCP Required - {row[20].value}\n"\
                f"Referral Required - {row[20].value}\n"\
                f"Patient eligible for WWE? - (Y/N)\n"\
                f"Covered at 100%\n"\
                f"Last WWE visit:-\n"\
                f"Spoke with - \n"\
                f"Ref# - \n"\
                f"Verified Through TriZetto on {today_date}"
            print('WWE_and_NP_WWE Notes Updated')
    else: 
        if row[20].value == 'Yes':
            row[33].value = f"DOS - {Appoinment_date}\n"\
                f"{row[13].value}\n"\
                f"Effective Date - {row[18].value}\n"\
                f"Plan Type - {result}\n"\
                f"PCP Required - {row[20].value}\n"\
                f"Referral Required - {row[20].value}\n"\
                f"PCP: {row[22].value}\n"\
                f"{line_here}\n"\
                f"Copayment - {row[32].value}\n"\
                f"{Ind_Ded}"\
                f"{Fam_Ded}"\
                f"{Ind_OOP}"\
                f"{Fam_OOP}"\
                f"Coinsurance - {row[31].value}\n"\
                f"Verified Through TriZetto on {today_date}"
            print('Commercial Notes Updated')
        else:
            row[33].value = f"DOS - {Appoinment_date}\n"\
                f"{row[13].value}\n"\
                f"Effective Date - {row[18].value}\n"\
                f"Plan Type - {result}\n"\
                f"PCP Required - {row[20].value}\n"\
                f"Referral Required - {row[20].value}\n"\
                f"{line_here}\n"\
                f"Copayment - {row[32].value}\n"\
                f"{Ind_Ded}"\
                f"{Fam_Ded}"\
                f"{Ind_OOP}"\
                f"{Fam_OOP}"\
                f"Coinsurance - {row[31].value}\n"\
                f"Verified Through TriZetto on {today_date}"
            print('Commercial Notes Updated')

def Medicare_Notes():
    if row[35].value == 'No' and row[36].value == 'NA':
        row[33].value = f"DOS - {Appoinment_date}\n"\
            f"{row[13].value}\n"\
            f"Effective Date - {row[18].value}\n"\
            f"Deductible Remaining - {row[24].value}\n"\
            f"Coinsurance - {row[31].value}\n"\
            f"Verified Through TriZetto on {today_date}"
        print('Medicare Notes Updated')


def ServiceTypeCode():
    try:
        professinal_office_98 = driver.find_element(By.XPATH,'//*[@id="plus_98 - Professional (Physician) Visit - Office"]').click()
    except:
        print('professinal_office_98 already selected')


def Mary_Santiago():
    try:
        select_provider = driver.find_element(By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')
        Mary_Santiago = Select(select_provider)
        Mary_Santiago.select_by_visible_text('Mary Santiago - 1528097268')
        print('Mary Santiago selected successfully')
    except:
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="btnEditProviderButton"]/span/span')))
        time.sleep(1.5)
        
        Edit_Provider = driver.find_element(By.XPATH,'//*[@id="btnEditProviderButton"]/span/span').click()
        time.sleep(2)
        Add_Provider = driver.find_element(By.XPATH,'//*[@id="Provider"]/div/a').click()
        time.sleep(1)
        provider_F_name = driver.find_element(By.XPATH,'//*[@id="ProviderFirstName"]').send_keys('Mary')
        provider_L_name = driver.find_element(By.XPATH,'//*[@id="ProviderLastName"]').send_keys('Santiago')
        provider_NPI = driver.find_element(By.XPATH,'//*[@id="NPI"]').send_keys('1528097268')
        time.sleep(1)
        provider_save = driver.find_element(By.XPATH,'//*[@id="Provider"]//a[text()="Save"]').click()
        BacktoRequest = driver.find_element(By.XPATH,'//*[@id="backtoddelink"]').click()
        time.sleep(3)
        print('Mary Santiago - 1528097268 Added Successfully')

        select_provider = driver.find_element(By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')
        Mary_Santiago = Select(select_provider)
        Mary_Santiago.select_by_visible_text('Mary Santiago - 1528097268')
        print('Mary Santiago selected successfully')

def Heidi_Schultz():
    Medicare_provider = driver.find_element(By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')
    Heidi_Schultz = Select(Medicare_provider)
    Heidi_Schultz.select_by_visible_text('Heidi Schultz MD PA - 1659309441')
    print('Heidi Schultz MD PA selected successfully')

count = 1
for row in sheet.iter_rows(min_row=2):
    
    if row[17].value == None and row[1].value != None:

        eligibility_inquiry = driver.find_element(By.XPATH,'//*[@id="Span1"]/p[1]/a')
        eligibility_inquiry.click()
        print('eligibility inquiry click Successfull')
        time.sleep(1)

        Primary_Insurance_Name = str(row[13].value)

        if Primary_Insurance_Name.startswith('BCBS'):
            BCBS_select = driver.find_element(By.XPATH,'//*[@id="Blue Cross Blue Shield"]/a').click()
            time.sleep(1)
            BCBS_Texas = driver.find_element(By.XPATH,'//*[@id="Blue Cross Blue Shield"]/ul/li/a[text()="BCBS Texas"]').click()
            print('PCBS Texas selected Successfully')
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')))
            time.sleep(1.5)
            Mary_Santiago()
            ServiceTypeCode()

        elif Primary_Insurance_Name.startswith('Medicare') or Primary_Insurance_Name.startswith('MEDICARE'):
            Medicare_select = driver.find_element(By.XPATH,'//*[@id="Medicare"]/a').click()
            time.sleep(1)
            Medicare = driver.find_element(By.XPATH,'//*[@id="Medicare"]/ul/li/a[text()="Medicare"]').click()
            print('Medicare selected Successfully')
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')))
            time.sleep(1.5)
            Heidi_Schultz()

        elif Primary_Insurance_Name.startswith('Marpai') or Primary_Insurance_Name.startswith('Aetna') or Primary_Insurance_Name.startswith('aetna'):
            Commercial_select = driver.find_element(By.XPATH,'//*[@id="Commercial"]/a').click()
            time.sleep(1)
            Aetna = driver.find_element(By.XPATH,'//*[@id="Commercial"]/ul/li/a[text()="Aetna"]').click()
            print('Aetna selected Successfully')
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')))
            time.sleep(1.5)
            Mary_Santiago()
            ServiceTypeCode()

        elif Primary_Insurance_Name.startswith('Humana'):
            Commercial_select = driver.find_element(By.XPATH,'//*[@id="Commercial"]/a').click()
            time.sleep(1)
            Humana = driver.find_element(By.XPATH,'//*[@id="Commercial"]/ul/li/a[text()="Humana"]').click()
            print('Humana selected Successfully')
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')))
            time.sleep(1.5)
            Mary_Santiago()
            ServiceTypeCode()

        elif Primary_Insurance_Name.startswith('Cigna') or Primary_Insurance_Name == 'Allegiance-(Cigna)' or Primary_Insurance_Name == '90 Degree Benefits (Cigna)':
            Commercial_select = driver.find_element(By.XPATH,'//*[@id="Commercial"]/a').click()
            time.sleep(1)
            Cigna = driver.find_element(By.XPATH,'//*[@id="Commercial"]/ul/li/a[text()="Cigna"]').click()
            print('Cigna selected Successfully')
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')))
            time.sleep(1.5)
            Mary_Santiago()
            ServiceTypeCode()

        elif Primary_Insurance_Name.startswith('AARP') or Primary_Insurance_Name.startswith('UHC') or Primary_Insurance_Name.startswith('United'):
            Commercial_select = driver.find_element(By.XPATH,'//*[@id="Commercial"]/a').click()
            time.sleep(1)
            UnitedHealthCare = driver.find_element(By.XPATH,'//*[@id="Commercial"]/ul/li/a[text()="UHC"]').click()
            print('United Health Care selected Successfully')
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')))
            time.sleep(1.5)
            Mary_Santiago()
            ServiceTypeCode()

        elif Primary_Insurance_Name.startswith('All Savers'):
            Commercial_select = driver.find_element(By.XPATH,'//*[@id="Commercial"]/a').click()
            time.sleep(1)
            All_savers = driver.find_element(By.XPATH,'//*[@id="Commercial"]/ul/li/a[text()="All Savers Insurance Eligibility Only"]').click()
            print('All Savers Insurance Eligibility Only selected Successfully')
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')))
            time.sleep(1.5)
            Mary_Santiago()
            ServiceTypeCode()

        elif Primary_Insurance_Name.startswith('Amerigroup'):
            Medicaid_select = driver.find_element(By.XPATH,'//*[@id="Medicaid"]/a').click()
            time.sleep(1)
            Amerigroup = driver.find_element(By.XPATH,'//*[@id="Medicaid"]/ul/li/a[text()="Amerigroup"]').click()
            print('Amerigroup selected Successfully')
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')))
            time.sleep(1.5)
            Mary_Santiago()
            ServiceTypeCode()

        elif Primary_Insurance_Name.startswith('Community'):
            Commercial_select = driver.find_element(By.XPATH,'//*[@id="Commercial"]/a').click()
            time.sleep(1)
            Community = driver.find_element(By.XPATH,'//*[@id="Commercial"]/ul/li/a[text()="Community Health Choice "]').click()
            print('Community Health Choice selected Successfully')
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')))
            time.sleep(1.5)
            Mary_Santiago()
            ServiceTypeCode()

        elif Primary_Insurance_Name.startswith('Golden'):
            Commercial_select = driver.find_element(By.XPATH,'//*[@id="Commercial"]/a').click()
            time.sleep(1)
            Golden_Rule = driver.find_element(By.XPATH,'//*[@id="Commercial"]/ul/li/a[text()="Golden Rule"]').click()
            print('Golden_Rule selected Successfully')
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')))
            time.sleep(1.5)
            Mary_Santiago()
            ServiceTypeCode()

        elif Primary_Insurance_Name.startswith('Meritain'):
            Commercial_select = driver.find_element(By.XPATH,'//*[@id="Commercial"]/a').click()
            time.sleep(1)
            Meritain_Health = driver.find_element(By.XPATH,'//*[@id="Commercial"]/ul/li/a[text()="Meritain Health"]').click()
            print(' Meritain_Health selected Successfully')
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')))
            time.sleep(1.5)
            Mary_Santiago()
            ServiceTypeCode()

        elif Primary_Insurance_Name.startswith('Tricare'):
            Military_select = driver.find_element(By.XPATH,'//*[@id="Military"]/a').click()
            time.sleep(1)
            Tricare_East = driver.find_element(By.XPATH,'//*[@id="Military"]/ul/li/a[text()="Tricare East"]').click()
            print('Tricare_East selected Successfully')
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')))
            time.sleep(1.5)
            Mary_Santiago()
            ServiceTypeCode()

        elif Primary_Insurance_Name.startswith('UMR') or Primary_Insurance_Name.startswith('US-UMR'):
            Commercial_select = driver.find_element(By.XPATH,'//*[@id="Commercial"]/a').click()
            time.sleep(1)
            UMR_Wausau = driver.find_element(By.XPATH,'//*[@id="Commercial"]/ul/li/a[text()="UMR-Wausau"]').click()
            print('UMR_Wausau selected Successfully')
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')))
            time.sleep(1.5)
            Mary_Santiago()
            ServiceTypeCode()

        elif Primary_Insurance_Name.startswith('Web'):
            Commercial_select = driver.find_element(By.XPATH,'//*[@id="Commercial"]/a').click()
            time.sleep(1)
            Web_TPA = driver.find_element(By.XPATH,'//*[@id="Commercial"]/ul/li/a[text()="Web TPA"]').click()
            print('Web_TPA selected Successfully')
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')))
            time.sleep(1.5)
            Mary_Santiago()
            ServiceTypeCode()

        elif Primary_Insurance_Name.startswith('WellCare') or Primary_Insurance_Name.startswith('Wellcare'):
            Commercial_select = driver.find_element(By.XPATH,'//*[@id="Commercial"]/a').click()
            time.sleep(1)
            WellCare = driver.find_element(By.XPATH,'//*[@id="Commercial"]/ul/li/a[text()="Wellcare/Harmony/Healthease/Staywell"]').click()
            print('Wellcare/Harmony/Healthease/Staywell selected Successfully')
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')))
            time.sleep(1.5)
            Mary_Santiago()
            ServiceTypeCode()

        elif Primary_Insurance_Name.startswith('Devoted') or Primary_Insurance_Name.startswith('DEVOTED'):
            Commercial_select = driver.find_element(By.XPATH,'//*[@id="Commercial"]/a').click()
            time.sleep(1)
            Devoted_health = driver.find_element(By.XPATH,'//*[@id="Commercial"]//li/a[text()="Devoted Health"]').click()
            print('Devoted Health selected Successfully')
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_SelectedProviderValue"]')))
            time.sleep(1.5)
            Mary_Santiago()
            ServiceTypeCode()
            
        else:
            print('Primary Insurance Name Not Available in Excel')
            row[17].value = 'Add in Bot'
            workbook.save(FFM_Bot_Output_File_Path)
            driver.back()
            time.sleep(1)
            continue

        excel_date = row[7].value
        if excel_date != None:
            Appoinment_date = excel_date.strftime("%m/%d/%Y")
            dateOfservice = driver.find_element(By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_DateOfService"]')
            dateOfservice.clear()
            dateOfservice.send_keys(Appoinment_date)
            try:
                DateofServiceEnd = driver.find_element(By.XPATH,'//*[@id="EligibilityRequestPayerInquiry_EligibilityRequestFieldValues_DateOfServiceEnd"]')
                DateofServiceEnd.clear()
                DateofServiceEnd.send_keys(Appoinment_date)
                print('Date of Service End click successfull')
            except:
                print('Date of Service End Not Present')
        else:
            print('None in Appointment date')

        Subscriber_No = row[14].value
        if Subscriber_No != None:
            subscriberID = driver.find_element(By.XPATH,'//*[@id="EligibilityRequestTemplateInquiry_EligibilityRequestFieldValues_InsuranceNum"]')
            subscriberID.send_keys(Subscriber_No)
            print('subscriberID selected successfully')
        else:
            subscriberID = driver.find_element(By.XPATH,'//*[@id="EligibilityRequestTemplateInquiry_EligibilityRequestFieldValues_InsuranceNum"]')
            subscriberID.send_keys('None')
            print('None in Insurance Number')

        try:
            patient_F_name = row[2].value
            if patient_F_name != None:
                subscriber_FirstName = driver.find_element(By.XPATH,'//*[@id="EligibilityRequestTemplateInquiry_EligibilityRequestFieldValues_InsuredFirstName"]')
                subscriber_FirstName.send_keys(patient_F_name)
                print('subscriber first name sent successfull')
            else:
                subscriber_FirstName = driver.find_element(By.XPATH,'//*[@id="EligibilityRequestTemplateInquiry_EligibilityRequestFieldValues_InsuredFirstName"]')
                subscriber_FirstName.send_keys('None')
                print('None in First Name')

            patient_L_name = row[4].value
            if patient_L_name != None:
                subscriber_LastName = driver.find_element(By.XPATH,'//*[@id="EligibilityRequestTemplateInquiry_EligibilityRequestFieldValues_InsuredLastName"]')
                subscriber_LastName.send_keys(patient_L_name)
                print('subscriber last name sent successfull')
            else:
                subscriber_LastName = driver.find_element(By.XPATH,'//*[@id="EligibilityRequestTemplateInquiry_EligibilityRequestFieldValues_InsuredLastName"]')
                subscriber_LastName.send_keys('None')
                print('None in Last Name')
        except:
            print('No option to enter subscribers First Name and Last Name')

        format_DOB = row[6].value
        if format_DOB != None:
            patient_DOB = format_DOB.strftime("%m/%d/%Y")
            subscriber_DateOfBirth = driver.find_element(By.XPATH,'//*[@id="EligibilityRequestTemplateInquiry_EligibilityRequestFieldValues_InsuredDob"]')
            subscriber_DateOfBirth.send_keys(patient_DOB)
            print('subscriber date of birth sent successfull')
        else:
            print('None in Date of Birth')

        time.sleep(1)
        Submit_eligibility = driver.find_element(By.XPATH,'//*[@id="btnUploadButton"]/span')
        Submit_eligibility.click()
        print('submit eligibility click successfull')
        time.sleep(3)
        
        try:
            driver.switch_to.alert.accept()
        except:
            print('Pop Up Not Available')

        WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="btnEditEligibilityResponseButton"]')))
        time.sleep(1.5)

        try:
            Coverage_in_text = driver.find_element(By.XPATH,'//*[@id="trnEligibilityStatus"]').text

            row[40].value = Coverage_in_text

            if Coverage_in_text == 'Active Coverage':

                row[17].value = 'Yes'

                Benefit_Information = driver.find_element(By.XPATH,'//*[@id="tab_benefitinformation"]/strong')
                Benefit_Information.click()
                print('Clicked on Benefit_Information')
                time.sleep(1)

                Expand_All = driver.find_element(By.XPATH,'//*[@id="ExpandInfo"]/li/a')
                Expand_All.click()
                time.sleep(1)
                print('Expand all click successfull')

                Payer_Logo = driver.find_element(By.XPATH,'//*[@id="PayerLogo"]//following::h3').text
                print('Payer Name:',Payer_Logo)

                if Payer_Logo == 'BCBS Texas':
                    BCBS()
                    Commercial_Notes()
                elif Payer_Logo == 'Medicare':
                    MEDICARE()
                    Medicare_Notes()
                elif Payer_Logo == 'UHC':
                    United_Healthcare()
                    Commercial_Notes()
                elif Payer_Logo == 'Aetna':
                    Aetna_Data()
                    Commercial_Notes()
                else:
                    Common_Data()
                    Commercial_Notes()

                current_DateTime = dt.now()
                US_format = current_DateTime.strftime("%m/%d/%Y %H:%M:%S")
                row[39].value = US_format
                
                workbook.save(FFM_Bot_Output_File_Path)
                print('Data Updated in Excel successfully')
                time.sleep(1)

                current_window = driver.current_window_handle
                print('Original window:',current_window)

                print_page = driver.find_element(By.XPATH,'//*[@id="exportPrint"]')
                print_page.click()
                print('Print page click successfull')
                
                all_windows = driver.window_handles
                driver.switch_to.window(all_windows[1])
                print('Switched to New window:',all_windows[1])

                WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.TAG_NAME,'body')))
                time.sleep(1.5)

                main_root = driver.find_element(By.TAG_NAME,'print-preview-app').shadow_root
                sidebar_root = main_root.find_element(By.CSS_SELECTOR,'[id="sidebar"]').shadow_root
                
                if count == 1:
                    container = sidebar_root.find_element(By.CSS_SELECTOR,'[id="container"]')
                    destinationSettings_root = shadadow_root3 = container.find_element(By.CSS_SELECTOR,'[id="destinationSettings"]').shadow_root
                    destinationSelect_root = destinationSettings_root.find_element(By.CSS_SELECTOR,'[id="destinationSelect"]').shadow_root
                    destinationDropdown = destinationSelect_root.find_element(By.CSS_SELECTOR,'[class="md-select"]')
                    Destination = Select(destinationDropdown)
                    Destination.select_by_visible_text("Save as PDF")
                    time.sleep(2)

                    more_settings_root = container.find_element(By.TAG_NAME,'print-preview-more-settings').shadow_root
                    more_settings_Dropdown = more_settings_root.find_element(By.CSS_SELECTOR,'[aria-label="More settings"]')
                    more_settings_Dropdown.click()
                    time.sleep(2)

                    moreSettings = container.find_element(By.CSS_SELECTOR,'[id="moreSettings"]')
                    options_settings_root = moreSettings.find_element(By.TAG_NAME,'print-preview-other-options-settings').shadow_root
                    headerFooter_root = options_settings_root.find_element(By.CSS_SELECTOR,'[id="headerFooter"]').shadow_root
                    cssBackground = options_settings_root.find_element(By.CSS_SELECTOR,'[id="cssBackground"]').shadow_root
                    CheckBox1 = headerFooter_root.find_element(By.CSS_SELECTOR,'[id="checkbox"]')
                    CheckBox2 = cssBackground.find_element(By.CSS_SELECTOR,'[id="checkbox"]')
                    driver.execute_script("arguments[0].scrollIntoView(true);", CheckBox1)
                    time.sleep(2)
                    CheckBox1.click()
                    CheckBox2.click()
                    time.sleep(3)

                time.sleep(2)
                strip_root = sidebar_root.find_element(By.CSS_SELECTOR,'print-preview-button-strip').shadow_root
                Save_Button = strip_root.find_element(By.CSS_SELECTOR,'[class="action-button"]')
                Save_Button.click()

                patientID = row[1].value
                APTdate = row[7].value
                PDF_name = str(patientID) +'_'+ str(APTdate).replace('00:00:00','')
                
                chrome_app = Application(backend='uia').connect(title_re="TriZetto Provider Solutions™ : - Google Chrome", timeout=50)
                window_switch = chrome_app.TrizettoProviderSolutionsGoogleChrome.child_window(title="Save As", control_type="Window")
                # chrome_app.TrizettoProviderSolutionsGoogleChrome.print_control_identifiers()

                keyboard.send_keys(f'{PDF_name}')
                time.sleep(1)

                if count == 1:
                    Previous_Locations = chrome_app.TrizettoProviderSolutionsGoogleChrome.child_window(title="Previous Locations", control_type="Button")
                    Previous_Locations.click_input()

                    keyboard.send_keys(today_folder_path, with_spaces=True)
                    keyboard.send_keys('{ENTER}')
                    time.sleep(1)

                Save = chrome_app.TrizettoProviderSolutionsGoogleChrome.child_window(title="Save", auto_id="1", control_type="Button")
                Save.click_input()

                try:
                    Replace_Yes = chrome_app.TrizettoProviderSolutionsGoogleChrome.child_window(title="Yes", auto_id="CommandButton_6", control_type="Button")
                    Replace_Yes.click_input()
                except:
                    pass

                time.sleep(3)
                print('PDF Downloaded Successfully')

                count+=1
                
                driver.switch_to.window(current_window)
                print('Switched back to the Original window:',current_window)
                driver.back()
                time.sleep(1)

            elif Coverage_in_text == 'Rejected':

                row[17].value = 'No'

                try:
                    eligErrorMsg = driver.find_elements(By.XPATH,'//*[@id="eligErrorMsg"]/span')
                    if len(eligErrorMsg) >= 2:
                        row[41].value = f"{eligErrorMsg[0].text}\n"\
                            f"{eligErrorMsg[1].text}"
                    else:
                        row[41].value = eligErrorMsg[0].text

                except:
                    row[41].value = 'Reason Not Available'

                workbook.save(FFM_Bot_Output_File_Path)
                print('Response for Rejected')
                time.sleep(2)
                driver.back()
                time.sleep(1)

            else:
                row[17].value = 'No'
                workbook.save(FFM_Bot_Output_File_Path)
                print('Response for Inactive')
                time.sleep(2)
                driver.back()
                time.sleep(1)
                
        except:
            row[17].value = 'No'
            row[40].value = 'Not Available'
            Request_to_Response = driver.find_element(By.XPATH,'//*[@id="eligibilityRequestResponse"]/div[3]/div[2]/h2').text
            row[41].value = Request_to_Response
            workbook.save(FFM_Bot_Output_File_Path)
            time.sleep(2)
            driver.back()
            time.sleep(1)
    else:
        continue

Active = 0
Inactive = 0
Rejected = 0
Response = 0
AddinBot = 0
Resubmit = 0

for Row in sheet.iter_rows(min_row=2):
    
    if Row[0].value != None:

        if Row[40].value == 'Active Coverage':
            Active += 1

        elif Row[40].value == 'Inactive':
            Inactive += 1

        elif Row[40].value == 'Rejected':
            Rejected += 1
        
        elif Row[40].value == 'No response':
            Response += 1

        elif Row[40].value == None:
            AddinBot += 1

        else:
            Resubmit += 1

SumofAll = Active + Inactive + Rejected + Response + AddinBot + Resubmit

print('Active:', Active)
print('Inactive:', Inactive)
print('Rejected:', Rejected)
print('No Response:', Response)
print('Add in Bot:', AddinBot)
print('Correct and Resubmit:', Resubmit)
print('Sum of All Total:', SumofAll)

workbook.close()
driver.close()
print('FFM Eligibility Bot Exicuted Successfully')

sender = 'sender@gmail.com'
to = ['random@gmail.com', 'random1@gmail.com']
cc = ['unknown@gmail.com', 'unknown1@gmail.com']

message = f"""From: {sender}
To: {to}
CC: {cc}
Subject: Eligibility and Benefits BOT Status: {today_date}
Hello Sir,

The Eligibility and Benefits BOT has been successfully completed.

Total Active: {Active}
Total Inactive: {Inactive}
Total Rejected: {Rejected}
Total No Response: {Response}
Total Add in Bot: {AddinBot}
Total Resubmit: {Resubmit}
Total Count: {SumofAll}

Please find the bot output file on below path:-
H:\AR PRODUCTION REPORTS\Business Intelligence\Eligibility and Benefits BOT Output.xlsx

Thanks & Regards,
Pawan Channe | Python Developer
Email: sender@gmail.com
Website: www.nath-mds.com """

with smtplib.SMTP('10.200.80.35', "25") as server:
    server.sendmail(sender, to + cc, message)
    print("The email has been sent successfully")

